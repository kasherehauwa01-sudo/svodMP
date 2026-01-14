from __future__ import annotations

import dataclasses
import datetime
import logging
import math
import re
from pathlib import Path
from typing import Any, Iterable, Optional

import openpyxl
import pandas as pd
import xlrd

logger = logging.getLogger(__name__)


@dataclasses.dataclass
class ExcelData:
    rows: list[list[Any]]
    header_row_index: int
    data_start_row: int
    data_end_row: int
    column_map: dict[str, int]
    date_col: int
    day_col: int


KEYWORDS = {
    "checks": "Чеки",
    "goods": "Товары",
    "gift_cert": "Подарочные сертификаты",
}
KEYWORD_ALIASES = {
    "goods": ["штуки"],
}
HEADER_ROWS = [2, 3, 4, 5]


class ExcelReadError(Exception):
    pass


class _DataFrameSheet:
    """Обертка над DataFrame для совместимости с xlrd-интерфейсом."""

    def __init__(self, rows: list[list[Any]]) -> None:
        self._rows = rows
        self.nrows = len(rows)
        self.ncols = max((len(row) for row in rows), default=0)
        self.merged_cells: list[tuple[int, int, int, int]] = []

    def cell_value(self, row: int, col: int) -> Any:
        try:
            value = self._rows[row][col]
        except IndexError:
            return None
        if isinstance(value, float) and math.isnan(value):
            return None
        return value


def read_excel_smart(path: Path) -> pd.DataFrame:
    """Читает Excel с явным движком для .xls."""
    if path.suffix.lower() == ".xls":
        return pd.read_excel(path, engine="xlrd")
    return pd.read_excel(path)


def _convert_xls_to_xlsx(file_path: Path) -> Path:
    """Конвертирует xls в xlsx для повторной попытки чтения."""
    workbook = xlrd.open_workbook(file_path)
    sheet = workbook.sheet_by_index(0)
    target_path = file_path.with_name(f"{file_path.stem}_copy.xlsx")
    output = openpyxl.Workbook()
    out_sheet = output.active
    out_sheet.title = sheet.name
    for row in range(sheet.nrows):
        for col in range(sheet.ncols):
            out_sheet.cell(row=row + 1, column=col + 1, value=sheet.cell_value(row, col))
    output.save(target_path)
    return target_path


def read_excel(file_path: Path) -> ExcelData:
    if file_path.suffix.lower() == ".xlsx":
        return _read_xlsx(file_path)
    if file_path.suffix.lower() == ".xls":
        try:
            return _read_xls(file_path)
        except ExcelReadError as exc:
            if "Не найдена строка" in str(exc):
                converted = _convert_xls_to_xlsx(file_path)
                logger.info("Файл %s конвертирован в %s", file_path.name, converted)
                return _read_xlsx(converted)
            raise
    raise ExcelReadError(f"Неподдерживаемое расширение: {file_path.suffix}")


def _read_xlsx(file_path: Path) -> ExcelData:
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = _select_sheet_for_parsing_xlsx(workbook)

    store = _detect_store_from_path(file_path)
    data_start_row, date_col, day_col = _find_data_start_row_xlsx(sheet)
    try:
        checks_header_row, _checks_header_col = _find_checks_header_cell_xlsx(sheet)
        data_start_row = checks_header_row + 5
    except ExcelReadError:
        logger.warning("Не найдена ячейка «Чеки» в заголовке, используем строку «Дата»")
    if store == "Ахтубинск":
        data_start_row = 7
    header_rows = _build_header_rows(data_start_row)
    header_row_index = header_rows[0] if header_rows else HEADER_ROWS[0]
    column_map = _find_keyword_columns_xlsx(sheet, header_rows or HEADER_ROWS, store)
    data_end_row = _find_data_end_row_xlsx(sheet, data_start_row)

    rows = _extract_rows_xlsx(sheet, data_start_row, data_end_row, column_map, date_col, day_col)

    return ExcelData(
        rows=rows,
        header_row_index=header_row_index,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
        column_map=column_map,
        date_col=date_col,
        day_col=day_col,
    )


def _read_xls(file_path: Path) -> ExcelData:
    sheet = _load_primary_xls_sheet(file_path)

    store = _detect_store_from_path(file_path)
    data_start_row, date_col, day_col = _find_data_start_row_xls(sheet)
    try:
        checks_header_row, _checks_header_col = _find_checks_header_cell_xls(sheet)
        data_start_row = checks_header_row + 6
    except ExcelReadError:
        logger.warning("Не найдена ячейка «Чеки» в заголовке, используем строку «Дата»")
    if store == "Ахтубинск":
        data_start_row = 7
    header_rows = _build_header_rows(data_start_row)
    header_row_index = header_rows[0] if header_rows else HEADER_ROWS[0]
    column_map = _find_keyword_columns_xls(
        sheet,
        header_rows or HEADER_ROWS,
        data_start_row,
        store,
    )
    data_end_row = _find_data_end_row_xls(sheet, data_start_row)

    rows = _extract_rows_xls(sheet, data_start_row, data_end_row, column_map, date_col, day_col)

    return ExcelData(
        rows=rows,
        header_row_index=header_row_index,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
        column_map=column_map,
        date_col=date_col,
        day_col=day_col,
    )


def _find_keyword_columns_xlsx(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    header_rows: list[int],
    store: str | None = None,
) -> dict[str, int]:
    column_map: dict[str, int] = {}
    try:
        column_map["checks"] = _find_header_column_xlsx(sheet, "checks")
    except ExcelReadError:
        column_map["checks"] = _get_store_fallback_column(store, "checks")

    try:
        column_map["goods"] = _find_header_column_xlsx(sheet, "goods")
    except ExcelReadError:
        column_map["goods"] = _get_store_fallback_column(store, "goods")

    try:
        column_map["gift_cert"] = _find_header_column_xlsx(sheet, "gift_cert")
    except ExcelReadError:
        column_map["gift_cert"] = _get_store_fallback_column(store, "gift_cert")

    _validate_column_map(column_map, header_rows)
    return column_map


def _get_merge_left_col_xlsx(
    sheet: openpyxl.worksheet.worksheet.Worksheet, row: int, col: int
) -> int:
    for cell_range in sheet.merged_cells.ranges:
        if cell_range.min_row <= row <= cell_range.max_row and cell_range.min_col <= col <= cell_range.max_col:
            return cell_range.min_col
    return col


def _find_keyword_columns_xls(
    sheet: xlrd.sheet.Sheet,
    header_rows: list[int],
    data_start_row: int,
    store: str | None = None,
) -> dict[str, int]:
    column_map: dict[str, int] = {}
    try:
        column_map["checks"] = _find_header_column_xls(sheet, "checks")
    except ExcelReadError:
        column_map["checks"] = _get_store_fallback_column(store, "checks")

    try:
        column_map["goods"] = _find_header_column_xls(sheet, "goods")
    except ExcelReadError:
        column_map["goods"] = _get_store_fallback_column(store, "goods")

    try:
        column_map["gift_cert"] = _find_header_column_xls(sheet, "gift_cert")
    except ExcelReadError:
        column_map["gift_cert"] = _get_store_fallback_column(store, "gift_cert")

    _validate_column_map(column_map, header_rows)
    return column_map


def _get_merge_left_col_xls(sheet: xlrd.sheet.Sheet, row: int, col: int) -> int:
    for rlo, rhi, clo, chi in sheet.merged_cells:
        if rlo <= row < rhi and clo <= col < chi:
            return clo
    return col


def _find_header_column_xlsx(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    key: str,
) -> int:
    """Возвращает индекс колонки заголовка в 3-й строке (xlsx)."""
    header_row = 3
    merged_col = _find_header_column_in_merged_xlsx(sheet, key, header_row)
    if merged_col is not None:
        return merged_col - 1
    max_col = sheet.max_column
    for col in range(1, max_col + 1):
        value = sheet.cell(row=header_row, column=col).value
        if _is_header_value(value, key):
            return col - 1
    raise ExcelReadError(f"Не найдена колонка «{KEYWORDS[key]}» в заголовке файла.")


def _find_header_column_xls(sheet: xlrd.sheet.Sheet, key: str) -> int:
    """Возвращает индекс колонки заголовка в 3-й строке (xls)."""
    header_row = 2
    merged_col = _find_header_column_in_merged_xls(sheet, key, header_row)
    if merged_col is not None:
        return merged_col
    for col in range(sheet.ncols):
        value = sheet.cell_value(header_row, col)
        if _is_header_value(value, key):
            return col
    raise ExcelReadError(f"Не найдена колонка «{KEYWORDS[key]}» в заголовке файла.")


def _find_header_column_in_merged_xlsx(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    key: str,
    header_row: int,
) -> Optional[int]:
    for cell_range in sheet.merged_cells.ranges:
        if cell_range.min_row != header_row:
            continue
        value = sheet.cell(row=cell_range.min_row, column=cell_range.min_col).value
        if _is_header_value(value, key):
            return cell_range.min_col
    return None


def _find_header_column_in_merged_xls(
    sheet: xlrd.sheet.Sheet,
    key: str,
    header_row: int,
) -> Optional[int]:
    for rlo, _rhi, clo, _chi in sheet.merged_cells:
        if rlo != header_row:
            continue
        value = sheet.cell_value(rlo, clo)
        if _is_header_value(value, key):
            return clo
    return None


def _is_header_value(value: Any, key: str) -> bool:
    text = _normalize_header_value(value)
    if text is None:
        return False
    if text == KEYWORDS[key].lower():
        return True
    aliases = KEYWORD_ALIASES.get(key, [])
    return any(text == alias.lower() for alias in aliases)


def _find_checks_header_cell_xlsx(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
) -> tuple[int, int]:
    """Возвращает координаты ячейки с заголовком «Чеки» (xlsx)."""
    for cell_range in sheet.merged_cells.ranges:
        value = sheet.cell(row=cell_range.min_row, column=cell_range.min_col).value
        if _is_header_value(value, "checks"):
            return cell_range.min_row, cell_range.min_col
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            value = sheet.cell(row=row, column=col).value
            if _is_header_value(value, "checks"):
                return row, col
    raise ExcelReadError("Не найдена колонка «Чеки» в заголовке файла.")


def _find_checks_header_cell_xls(sheet: xlrd.sheet.Sheet) -> tuple[int, int]:
    """Возвращает координаты ячейки с заголовком «Чеки» (xls)."""
    for rlo, _rhi, clo, _chi in sheet.merged_cells:
        value = sheet.cell_value(rlo, clo)
        if _is_header_value(value, "checks"):
            return rlo, clo
    for row in range(sheet.nrows):
        for col in range(sheet.ncols):
            value = sheet.cell_value(row, col)
            if _is_header_value(value, "checks"):
                return row, col
    raise ExcelReadError("Не найдена колонка «Чеки» в заголовке файла.")


def _read_xls_merged_cells(file_path: Path) -> list[tuple[int, int, int, int]]:
    """Считывает диапазоны объединённых ячеек из .xls."""
    try:
        workbook = xlrd.open_workbook(file_path, formatting_info=True)
        sheet = workbook.sheet_by_index(0)
        return list(sheet.merged_cells)
    except Exception as exc:  # noqa: BLE001 - резервный путь, если формат не поддержан.
        logger.warning("Не удалось считать объединённые ячейки из %s: %s", file_path.name, exc)
        return []


def _select_sheet_for_parsing_xlsx(
    workbook: openpyxl.Workbook,
) -> openpyxl.worksheet.worksheet.Worksheet:
    """Пытается найти лист с заголовками, иначе возвращает активный."""
    for sheet in workbook.worksheets:
        try:
            _find_checks_header_cell_xlsx(sheet)
            return sheet
        except ExcelReadError:
            continue
    return workbook.active


def _load_primary_xls_sheet(file_path: Path) -> _DataFrameSheet:
    """Загружает .xls лист, подбирая подходящий по наличию заголовков."""
    workbook = xlrd.open_workbook(file_path, formatting_info=True)
    selected_rows: list[list[Any]] = []
    selected_merged: list[tuple[int, int, int, int]] = []
    for idx in range(workbook.nsheets):
        sheet = workbook.sheet_by_index(idx)
        rows = [
            [
                sheet.cell_value(row, col)
                for col in range(sheet.ncols)
            ]
            for row in range(sheet.nrows)
        ]
        wrapper = _DataFrameSheet(rows)
        wrapper.merged_cells = list(sheet.merged_cells)
        try:
            _find_checks_header_cell_xls(wrapper)
            return wrapper
        except ExcelReadError:
            if not selected_rows:
                selected_rows = rows
                selected_merged = list(sheet.merged_cells)
    wrapper = _DataFrameSheet(selected_rows)
    wrapper.merged_cells = selected_merged
    return wrapper


def _find_data_start_row_xlsx(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
) -> tuple[int, int, int]:
    for row in range(1, sheet.max_row + 1):
        value = sheet.cell(row=row, column=1).value
        if _is_date_header(value):
            return row + 1, 0, 1
    return 2, 0, 1


def _find_data_start_row_xls(sheet: xlrd.sheet.Sheet) -> tuple[int, int, int]:
    for row in range(sheet.nrows):
        value = _get_header_text_xls(sheet, row, 0)
        if _is_date_header(value):
            return row + 2, 0, 1
    return 2, 0, 1


def _detect_store_from_path(file_path: Path) -> str | None:
    lower_name = file_path.stem.lower()
    if "ахтубинск" in lower_name:
        return "Ахтубинск"
    if "европа" in lower_name:
        return "Европа"
    if "санвэй" in lower_name or "санвей" in lower_name:
        return "Козловская"
    return None


def _get_store_fallback_column(store: str | None, keyword: str) -> int:
    if store in {"Ахтубинск", "Европа"} and keyword == "checks":
        return 16
    if store == "Ахтубинск" and keyword == "goods":
        return 19
    if store == "Ахтубинск" and keyword == "gift_cert":
        return 38
    if store == "Европа" and keyword == "goods":
        return 19
    if store == "Козловская" and keyword == "checks":
        return 19
    if store == "Козловская" and keyword == "goods":
        return 22
    raise ExcelReadError(f"Не найдена колонка «{KEYWORDS[keyword]}» в заголовке файла.")


def _find_data_end_row_xlsx(sheet: openpyxl.worksheet.worksheet.Worksheet, start_row: int) -> int:
    last_row = sheet.max_row
    return _trim_trailing_empty_rows(
        range(start_row, last_row + 1),
        lambda r: _row_has_data_xlsx(sheet, r),
    )


def _find_data_end_row_xls(sheet: xlrd.sheet.Sheet, start_row: int) -> int:
    last_row = sheet.nrows
    return _trim_trailing_empty_rows(
        range(start_row, last_row + 1),
        lambda r: _row_has_data_xls(sheet, r - 1),
    )


def _trim_trailing_empty_rows(rows: Iterable[int], has_data) -> int:
    rows = list(rows)
    for row in reversed(rows):
        if has_data(row):
            return row
    return rows[0] - 1


def _row_has_data_xlsx(sheet: openpyxl.worksheet.worksheet.Worksheet, row: int) -> bool:
    for col in range(1, 9):
        value = sheet.cell(row=row, column=col).value
        if not _is_empty_value(value):
            return True
    return False


def _row_has_data_xls(sheet: xlrd.sheet.Sheet, row: int) -> bool:
    for col in range(8):
        value = sheet.cell_value(row, col)
        if not _is_empty_value(value):
            return True
    return False


def _extract_rows_xlsx(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int,
    end_row: int,
    column_map: dict[str, int],
    date_col: int,
    day_col: int,
) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row in range(start_row, end_row + 1):
        rows.append(_build_row_xlsx(sheet, row, column_map, date_col, day_col))
    return rows


def _extract_rows_xls(
    sheet: xlrd.sheet.Sheet,
    start_row: int,
    end_row: int,
    column_map: dict[str, int],
    date_col: int,
    day_col: int,
) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row in range(start_row - 1, end_row):
        rows.append(_build_row_xls(sheet, row, column_map, date_col, day_col))
    return rows


def _build_row_xlsx(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    column_map: dict[str, int],
    date_col: int,
    day_col: int,
) -> list[Any]:
    values = [
        sheet.cell(row=row, column=date_col + 1).value,
        sheet.cell(row=row, column=day_col + 1).value,
        sheet.cell(row=row, column=3).value,
        sheet.cell(row=row, column=column_map["checks"] + 1).value,
        None,
        sheet.cell(row=row, column=column_map["goods"] + 1).value,
        sheet.cell(row=row, column=5).value,
        sheet.cell(row=row, column=column_map["gift_cert"] + 1).value,
    ]
    return values


def _build_row_xls(
    sheet: xlrd.sheet.Sheet,
    row: int,
    column_map: dict[str, int],
    date_col: int,
    day_col: int,
) -> list[Any]:
    values = [
        sheet.cell_value(row, date_col),
        sheet.cell_value(row, day_col),
        sheet.cell_value(row, 2),
        sheet.cell_value(row, column_map["checks"]),
        None,
        sheet.cell_value(row, column_map["goods"]),
        sheet.cell_value(row, 4),
        sheet.cell_value(row, column_map["gift_cert"]),
    ]
    return values


def _validate_column_map(column_map: dict[str, int], header_rows: list[int]) -> None:
    missing = [keyword for keyword in KEYWORDS if keyword not in column_map]
    if missing:
        raise ExcelReadError(
            "Не найдены заголовки в строках "
            f"{', '.join(str(row) for row in header_rows)}: "
            + ", ".join(KEYWORDS[key] for key in missing)
        )


def _build_header_rows(data_start_row: int) -> list[int]:
    if data_start_row <= 1:
        return []
    return list(range(1, data_start_row))


def _normalize_header_value(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).replace("\u00a0", " ").strip().lower()
    text = " ".join(text.split())
    return text if text else None


def _keyword_in_text(key: str, keyword: str, text: str) -> bool:
    if keyword.lower() in text:
        return True
    aliases = KEYWORD_ALIASES.get(key, [])
    return any(alias.lower() in text for alias in aliases)


def _is_date_header(value: Any) -> bool:
    if value is None:
        return False
    text = str(value).strip().lower()
    text = _normalize_text_for_header(text)
    return "дата" in text


def _is_day_header(value: Any) -> bool:
    if value is None:
        return False
    text = str(value).strip().lower()
    text = _normalize_text_for_header(text)
    return "день нед" in text


def _normalize_text_for_header(text: str) -> str:
    text = text.replace("\u00a0", " ")
    text = text.replace("\ufeff", " ")
    text = re.sub(r"[^\w]+", " ", text, flags=re.UNICODE)
    return " ".join(text.split())


def _is_empty_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value == "":
        return True
    return False


def _get_header_text_xlsx(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    col: int,
) -> Optional[str]:
    value = sheet.cell(row=row, column=col).value
    text = _normalize_header_value(value)
    if text:
        return text
    for cell_range in sheet.merged_cells.ranges:
        if cell_range.min_row <= row <= cell_range.max_row and cell_range.min_col <= col <= cell_range.max_col:
            merged_value = sheet.cell(row=cell_range.min_row, column=cell_range.min_col).value
            return _normalize_header_value(merged_value)
    return None


def _get_header_text_xls(sheet: xlrd.sheet.Sheet, row: int, col: int) -> Optional[str]:
    value = sheet.cell_value(row, col)
    text = _normalize_header_value(value)
    if text:
        return text
    for rlo, rhi, clo, chi in sheet.merged_cells:
        if rlo <= row < rhi and clo <= col < chi:
            merged_value = sheet.cell_value(rlo, clo)
            return _normalize_header_value(merged_value)
    return None


def _find_date_like_row_xlsx(sheet: openpyxl.worksheet.worksheet.Worksheet) -> Optional[int]:
    max_row = sheet.max_row
    max_col = sheet.max_column
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            value = sheet.cell(row=row, column=col).value
            if _is_date_like_value(value):
                return row + 1
    return None


def _find_date_like_row_xls(sheet: xlrd.sheet.Sheet) -> Optional[int]:
    max_col = sheet.ncols
    for row in range(sheet.nrows):
        for col in range(max_col):
            value = sheet.cell_value(row, col)
            if _is_date_like_value(value):
                return row + 2
    return None


def _find_date_like_in_column_xlsx(sheet: openpyxl.worksheet.worksheet.Worksheet) -> Optional[int]:
    max_row = sheet.max_row
    for row in range(1, max_row + 1):
        value = sheet.cell(row=row, column=1).value
        if _is_date_like_value(value):
            return row
    return None


def _find_date_like_in_column_xls(sheet: xlrd.sheet.Sheet) -> Optional[int]:
    for row in range(sheet.nrows):
        value = sheet.cell_value(row, 0)
        if _is_date_like_value(value):
            return row + 1
    return None


def _is_date_like_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, datetime.datetime):
        return True
    if isinstance(value, datetime.date):
        return True
    text = str(value).strip()
    if not text:
        return False
    for fmt in ("%d.%m.%y", "%d.%m.%Y"):
        try:
            datetime.datetime.strptime(text, fmt)
            return True
        except ValueError:
            continue
    return False


def _find_day_header_xlsx(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
) -> Optional[tuple[int, int]]:
    for row in range(1, sheet.max_row + 1):
        for col in range(1, min(sheet.max_column, 26) + 1):
            value = _get_header_text_xlsx(sheet, row, col)
            if _is_day_header(value):
                return row, col
    return None


def _find_day_header_xls(sheet: xlrd.sheet.Sheet) -> Optional[tuple[int, int]]:
    for row in range(sheet.nrows):
        for col in range(min(sheet.ncols, 26)):
            value = _get_header_text_xls(sheet, row, col)
            if _is_day_header(value):
                return row, col
    return None
